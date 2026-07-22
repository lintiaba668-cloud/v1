# -*- coding: utf-8 -*-
"""
Excel loader for project import.

Read project code and project name from xlsx files.
"""

from pathlib import Path


class ExcelLoader:

    def load(self, file_path):
        path = Path(file_path)

        if path.suffix.lower() != '.xlsx':
            raise ValueError('仅支持xlsx文件')

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError('缺少openpyxl依赖')

        workbook = load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True
        )

        sheet = workbook.active
        rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            code = row[0] if len(row) > 0 else ''
            name = row[1] if len(row) > 1 else ''

            rows.append([
                code,
                name
            ])

        workbook.close()

        return rows
